import streamlit as st
import openpyxl as ox
import pandas as pd
import datetime

# データを読み込み、worksheet型と1次元dfのリストにする デッキリストだけは別でリストに


def datastopy(path):
    dueldatas_master = ox.load_workbook(path)
    dueldatas = dueldatas_master["シート1"]
    datas = []
    decks = []
    i = 7
    while True:
        if dueldatas.cell(row=i, column=1).value is None:
            break
        datefront = dueldatas.cell(row=i, column=1).value
        decks.append(dueldatas.cell(row=i, column=2).value)
        duel = []
        for j in range(2, 15):
            duel.append(dueldatas.cell(row=i, column=j).value)
        dfduel = pd.DataFrame(
            {
                "デッキ": duel[0],
                "対戦数": duel[1],
                "先手": duel[2],
                "後手": duel[3],
                "先手勝ち": duel[4],
                "先手負け": duel[5],
                "後手勝ち": duel[6],
                "後手負け": duel[7],
                "先手勝率": pd.Series(duel[8], dtype=float),
                "後手勝率": pd.Series(duel[9], dtype=float),
                "勝ち": duel[10],
                "負け": duel[11],
                "勝率": pd.Series(duel[12], dtype=float),
            },
            index=[datefront],
        )
        datas.append(dfduel)
        i += 1
    return dueldatas_master, datas, decks


def pytodatas(dueldatas_master, deck, order, result):
    import datetime

    dueldatas = dueldatas_master["シート1"]
    datefront = datetime.datetime.now()
    datefront = "{}/{}/{}".format(datefront.year, datefront.month, datefront.day)
    i = 7
    cal = 1
    while True:
        dueldate = dueldatas.cell(row=i, column=1).value
        deckdataed = dueldatas.cell(row=i, column=2).value
        if dueldate is None:
            cal = 0
            break
        if dueldate == datefront and deckdataed == deck:
            if order == "先手" and result == "勝ち":
                dueldatas.cell(row=i, column=1, value=datefront)
                changepoint = int(dueldatas.cell(row=i, column=3).value)
                dueldatas.cell(row=i, column=3, value=changepoint + 1)
                changepoint = int(dueldatas.cell(row=i, column=4).value)
                dueldatas.cell(row=i, column=4, value=changepoint + 1)
                changepoint = int(dueldatas.cell(row=i, column=6).value)
                dueldatas.cell(row=i, column=6, value=changepoint + 1)
                break
            elif order == "後手" and result == "勝ち":
                dueldatas.cell(row=i, column=1, value=datefront)
                changepoint = int(dueldatas.cell(row=i, column=3).value)
                dueldatas.cell(row=i, column=3, value=changepoint + 1)
                changepoint = int(dueldatas.cell(row=i, column=5).value)
                dueldatas.cell(row=i, column=5, value=changepoint + 1)
                changepoint = int(dueldatas.cell(row=i, column=8).value)
                dueldatas.cell(row=i, column=8, value=changepoint + 1)
                break
            elif order == "先手" and result == "負け":
                dueldatas.cell(row=i, column=1, value=datefront)
                changepoint = int(dueldatas.cell(row=i, column=3).value)
                dueldatas.cell(row=i, column=3, value=changepoint + 1)
                changepoint = int(dueldatas.cell(row=i, column=4).value)
                dueldatas.cell(row=i, column=4, value=changepoint + 1)
                changepoint = int(dueldatas.cell(row=i, column=7).value)
                dueldatas.cell(row=i, column=7, value=changepoint + 1)
                break
            elif order == "後手" and result == "負け":
                dueldatas.cell(row=i, column=1, value=datefront)
                changepoint = int(dueldatas.cell(row=i, column=3).value)
                dueldatas.cell(row=i, column=3, value=changepoint + 1)
                changepoint = int(dueldatas.cell(row=i, column=5).value)
                dueldatas.cell(row=i, column=5, value=changepoint + 1)
                changepoint = int(dueldatas.cell(row=i, column=9).value)
                dueldatas.cell(row=i, column=9, value=changepoint + 1)
                break
        i += 1
    if cal == 0:
        if order == "先手" and result == "勝ち":
            dueldatas.cell(row=i, column=1, value=datefront)
            dueldatas.cell(row=i, column=2, value=deck)
            dueldatas.cell(row=i, column=3, value=1)
            dueldatas.cell(row=i, column=4, value=1)
            dueldatas.cell(row=i, column=5, value=0)
            dueldatas.cell(row=i, column=6, value=1)
            dueldatas.cell(row=i, column=7, value=0)
            dueldatas.cell(row=i, column=8, value=0)
            dueldatas.cell(row=i, column=9, value=0)

        elif order == "後手" and result == "勝ち":
            dueldatas.cell(row=i, column=1, value=datefront)
            dueldatas.cell(row=i, column=2, value=deck)
            dueldatas.cell(row=i, column=3, value=1)
            dueldatas.cell(row=i, column=4, value=0)
            dueldatas.cell(row=i, column=5, value=1)
            dueldatas.cell(row=i, column=6, value=0)
            dueldatas.cell(row=i, column=7, value=0)
            dueldatas.cell(row=i, column=8, value=1)
            dueldatas.cell(row=i, column=9, value=0)

        elif order == "先手" and result == "負け":
            dueldatas.cell(row=i, column=1, value=datefront)
            dueldatas.cell(row=i, column=2, value=deck)
            dueldatas.cell(row=i, column=3, value=1)
            dueldatas.cell(row=i, column=4, value=1)
            dueldatas.cell(row=i, column=5, value=0)
            dueldatas.cell(row=i, column=6, value=0)
            dueldatas.cell(row=i, column=7, value=1)
            dueldatas.cell(row=i, column=8, value=0)
            dueldatas.cell(row=i, column=9, value=0)

        elif order == "後手" and result == "負け":
            dueldatas.cell(row=i, column=1, value=datefront)
            dueldatas.cell(row=i, column=2, value=deck)
            dueldatas.cell(row=i, column=3, value=1)
            dueldatas.cell(row=i, column=4, value=0)
            dueldatas.cell(row=i, column=5, value=1)
            dueldatas.cell(row=i, column=6, value=0)
            dueldatas.cell(row=i, column=7, value=0)
            dueldatas.cell(row=i, column=8, value=0)
            dueldatas.cell(row=i, column=9, value=1)


def pytoadditionaldata(df, dueldatas):
    for i in range(len(df)):
        first = df.iloc[i]["先手"]
        firstwin = df.iloc[i]["先手勝ち"]
        second = df.iloc[i]["後手"]
        secondwin = df.iloc[i]["後手勝ち"]
        general = df.iloc[i]["対戦数"]
        generalwin = firstwin + secondwin
        if first != 0:
            rate_firstwin = round(firstwin / first, 3)
            dueldatas.cell(row=i + 7, column=10, value=rate_firstwin)
            df.iat[i, 8] = rate_firstwin
        if second != 0:
            rate_secondwin = round(secondwin / second, 3)
            dueldatas.cell(row=i + 7, column=11, value=rate_secondwin)
            df.iat[i, 9] = rate_secondwin
        if general != 0:
            rate_general = round(generalwin / general, 3)
            dueldatas.cell(row=i + 7, column=12, value=generalwin)
            dueldatas.cell(row=i + 7, column=13, value=general - generalwin)
            dueldatas.cell(row=i + 7, column=14, value=rate_general)
            df.iat[i, 10] = generalwin
            df.iat[i, 11] = general - generalwin
            df.iat[i, 12] = rate_general


def datas_init(dueldatas_master):
    dueldatas = dueldatas_master["シート1"]
    for row in dueldatas.iter_rows(min_row=7, min_col=1, max_row=600, max_col=11):
        for cell in row:
            cell.value = None


def advanceddata(df):
    import datetime

    today = datetime.datetime.now()
    today = "{}/{}/{}".format(today.year, today.month, today.day)
    sumfirst = 0
    sumfirstwin = 0
    sumsecond = 0
    sumsecondwin = 0
    sumduel = 0
    duelwin = 0
    for i in range(len(df)):
        addduelfirst = df.iloc[i]["先手"]
        addduelsecond = df.iloc[i]["後手"]
        addduelwinfirst = df.iloc[i]["先手勝ち"]
        addduelwinsecond = df.iloc[i]["後手勝ち"]
        sumfirst += addduelfirst
        sumsecond += addduelsecond
        sumfirstwin += addduelwinfirst
        sumsecondwin += addduelwinsecond
        sumduel += addduelfirst + addduelsecond
        duelwin += addduelwinfirst + addduelwinsecond
    winrate = None
    winratefirst = None
    winratesecond = None
    if sumduel != 0:
        winrate = round(duelwin / sumduel, 3)
    if sumfirst != 0:
        winratefirst = round(sumfirstwin / sumfirst, 3)
    if sumsecond != 0:
        winratesecond = round(sumsecondwin / sumsecond, 3)
    dfad = pd.DataFrame(
        {
            "総対戦数": sumduel,
            "全体勝率": winrate,
            "総勝ち数": duelwin,
            "総負け数": (sumduel - duelwin),
            "総先手数": sumfirst,
            "総後手数": sumsecond,
            "先手勝率": winratefirst,
            "後手勝率": winratesecond,
        },
        index=["{}現在".format(today)],
    )
    return dfad


# ページレイアウト
st.set_page_config(
    page_title="DC 戦績記入,分析ツール",
    page_icon="🧊",
    layout="wide",
    initial_sidebar_state="collapsed",
    menu_items={},
)

# 対戦データの読み込み、デッキ表示
dueldatas_master, datalist, deckdueled = datastopy("database_florting/dueldatas.xlsx")
dueldatas = dueldatas_master["シート1"]

# 　表示の調整　重複するデッキをはじいている
deckdueled = set(deckdueled)
deckdueled = list(deckdueled)

if "decks" not in st.session_state:
    st.session_state.decks = deckdueled

deck_options = st.session_state.decks


# ここまでが読み込み　ここから動的な部分

st.title("DC 戦績記入,分析ツール")

# デッキ情報の取り出し　記入モジュール
newdeck = st.text_input(
    "新しいデッキの追加 変なデッキを追加したら、戦績記入の前にリロードすること"
)
if st.button("追加"):
    apd = 0
    for i in st.session_state.decks:
        if newdeck == i:
            st.write("そのデッキは追加されています")
            apd = 1
            break
    if apd == 0 and newdeck != "":
        deck_options.append(newdeck)

# 選択肢に実質的なプレイスホルダーを追加
reject = "ここに入力 元の文字は消さない"
if deck_options == []:
    deck_options.append(reject)
else:
    if deck_options[0] != reject:
        deck_options.insert(0, reject)

deck = st.selectbox(
    "対戦したデッキを選んでください 直打ちで検索もできます", deck_options
)
order = st.radio("先手後手を記入", ("先手", "後手"), horizontal=True)
result = st.radio("勝ち負けを記入", ("勝ち", "負け"), horizontal=True)

submit = st.button("結果を記入")
# submit により書き込み
if submit is True:
    if deck == reject:
        st.write("無効なデッキ名です")
    else:
        pytodatas(dueldatas_master, deck, order, result)
        dueldatas_master.save("database_florting/dueldatas.xlsx")
        st.button("データの同期")

st.markdown("#### 対戦デッキ別データ")
if datalist == []:
    st.write("データがありません。")
    today = datetime.datetime.now()
    today = "{}/{}/{}".format(today.year, today.month, today.day)
    df = pd.DataFrame(
        {
            "デッキ": "",
            "対戦数": 0,
            "勝率": 0,
            "先手": 0,
            "後手": 0,
            "先手勝ち": 0,
            "先手負け": 0,
            "後手勝ち": 0,
            "後手負け": 0,
            "先手勝率": 0,
            "後手勝率": 0,
        },
        index=[today],
    )
else:
    df = pd.concat(
        datalist
    )  # ここについてエラーを吐かれる（空のデータフレームをconcatできません）が、そのような場合は条件分岐ではじいている
    pytoadditionaldata(df, dueldatas)
    dueldatas_master.save("database_florting/dueldatas.xlsx")
    st.write(df)

st.markdown("#### 全体データ")
dfad = advanceddata(df)
st.write(dfad)

st.markdown("### 危険　全データの初期化")
st.write(
    "仕様上,600種以上のデッキデータがある場合はexcelファイルから直接消去してください。"
)
st.write("その際はインデックス（見出し）まで消さないように")
check = st.checkbox("初期化しますか？")
check2 = st.checkbox("こうかいしませんね？")
if st.button("上の2つのチェック+このボタンでデータが初期化") and check and check2:
    datas_init(dueldatas_master)
    dueldatas_master.save("database_florting/dueldatas.xlsx")
    st.write("データを初期化しました。　リロードすると反映されます")
